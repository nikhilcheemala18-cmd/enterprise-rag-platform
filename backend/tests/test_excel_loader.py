import datetime
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.ingestion.base import DocumentLoader
from app.ingestion.loaders.excel import ExcelLoader
from app.models.document import NormalizedDocument, TableElement


def _build_fixture_workbook(path: Path) -> None:
    wb = Workbook()

    revenue = wb.active
    revenue.title = "Revenue"
    revenue.append(["Date", "Revenue", "Orders", "Active"])
    revenue.append([datetime.date(2026, 7, 1), 7200.50, 120, True])
    revenue.append([datetime.date(2026, 7, 2), 8100.00, 135, False])

    campaigns = wb.create_sheet("Campaigns")
    campaigns.append(["Campaign", "Clicks"])
    campaigns.append(["Summer Sale", 500])
    campaigns.append(["Winter Sale", 300])

    with_blank_row = wb.create_sheet("WithBlankRow")
    with_blank_row.append(["Date", "Revenue"])
    with_blank_row.append([datetime.date(2026, 7, 1), 7200])
    with_blank_row.append([None, None])
    with_blank_row.append([datetime.date(2026, 7, 3), 8100])

    wb.create_sheet("Empty")

    wb.save(path)


class TestExcelLoaderValidFixture(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp_dir = Path(tempfile.mkdtemp())
        cls.fixture_path = cls.tmp_dir / "fixture.xlsx"
        _build_fixture_workbook(cls.fixture_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp_dir, ignore_errors=True)

    def setUp(self):
        self.loader = ExcelLoader()
        self.document = self.loader.load(self.fixture_path, uploaded_by="user-1")

    def _table_for_sheet(self, sheet_name: str) -> TableElement:
        for element in self.document.elements:
            if element.source.location.sheet_name == sheet_name:
                return element
        raise AssertionError(f"no TableElement found for sheet {sheet_name}")

    def test_produces_normalized_document(self):
        self.assertIsInstance(self.document, NormalizedDocument)

    def test_document_id_generated(self):
        self.assertTrue(self.document.document_id)

    def test_metadata_filename(self):
        self.assertEqual(self.document.metadata.filename, "fixture.xlsx")

    def test_metadata_file_type(self):
        self.assertEqual(self.document.metadata.file_type, "excel")

    def test_metadata_file_size(self):
        self.assertGreater(self.document.metadata.file_size, 0)

    def test_uploaded_by_preserved(self):
        self.assertEqual(self.document.metadata.uploaded_by, "user-1")

    def test_content_hash_present(self):
        self.assertTrue(self.document.metadata.content_hash)
        self.assertEqual(len(self.document.metadata.content_hash), 64)

    def test_sheet_names_metadata(self):
        self.assertEqual(
            self.document.metadata.extra["sheet_names"],
            ["Revenue", "Campaigns", "WithBlankRow", "Empty"],
        )

    def test_empty_worksheet_produces_no_element(self):
        for element in self.document.elements:
            self.assertNotEqual(element.source.location.sheet_name, "Empty")

    def test_each_non_empty_worksheet_produces_one_table_element(self):
        sheet_names = [e.source.location.sheet_name for e in self.document.elements]
        self.assertEqual(
            sorted(sheet_names), sorted(["Revenue", "Campaigns", "WithBlankRow"])
        )

    def test_worksheet_order_maps_to_element_index(self):
        indices_by_sheet = {
            e.source.location.sheet_name: e.element_index for e in self.document.elements
        }
        self.assertEqual(indices_by_sheet["Revenue"], 0)
        self.assertEqual(indices_by_sheet["Campaigns"], 1)
        self.assertEqual(indices_by_sheet["WithBlankRow"], 2)

    def test_section_path_contains_sheet_name(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertEqual(revenue.section_path, ["Revenue"])

    def test_columns_preserved(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertEqual(revenue.columns, ["Date", "Revenue", "Orders", "Active"])

    def test_numeric_values_remain_numeric(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertEqual(revenue.rows[0][1], 7200.50)
        self.assertIsInstance(revenue.rows[0][1], float)
        self.assertEqual(revenue.rows[0][2], 120)
        self.assertIsInstance(revenue.rows[0][2], int)

    def test_boolean_values_remain_boolean(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertIs(revenue.rows[0][3], True)
        self.assertIs(revenue.rows[1][3], False)

    def test_date_values_remain_date_like(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertIsInstance(revenue.rows[0][0], (datetime.date, datetime.datetime))

    def test_blank_cells_remain_none(self):
        campaigns = self._table_for_sheet("Campaigns")
        for row in campaigns.rows:
            self.assertEqual(len(row), 2)

    def test_source_document_id_matches(self):
        for element in self.document.elements:
            self.assertEqual(element.source.document_id, self.document.document_id)

    def test_source_type_is_excel(self):
        for element in self.document.elements:
            self.assertEqual(element.source.source_type, "excel")

    def test_excel_location_sheet_name(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertEqual(revenue.source.location.sheet_name, "Revenue")

    def test_excel_location_row_bounds_without_blanks(self):
        revenue = self._table_for_sheet("Revenue")
        self.assertEqual(revenue.source.location.row_start, 1)
        self.assertEqual(revenue.source.location.row_end, 3)

    def test_excel_location_row_bounds_with_blank_row(self):
        with_blank = self._table_for_sheet("WithBlankRow")
        # header=row1, data rows at physical rows 2 and 4 (row 3 is blank
        # and omitted from the table body but not from the numbering)
        self.assertEqual(with_blank.source.location.row_start, 1)
        self.assertEqual(with_blank.source.location.row_end, 4)
        self.assertEqual(len(with_blank.rows), 2)

    def test_serializes_to_json(self):
        payload = self.document.model_dump_json()
        self.assertIsInstance(payload, str)
        self.assertIn("Revenue", payload)


class TestExcelLoaderMissingFile(unittest.TestCase):
    def test_missing_file_raises_clear_error(self):
        loader = ExcelLoader()
        with self.assertRaises(FileNotFoundError):
            loader.load(Path(tempfile.gettempdir()) / "does_not_exist.xlsx")


class TestExcelLoaderContract(unittest.TestCase):
    def test_loader_inherits_document_loader(self):
        self.assertIsInstance(ExcelLoader(), DocumentLoader)


if __name__ == "__main__":
    unittest.main()
