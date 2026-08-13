import hashlib
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.base import DocumentLoader
from app.models.document import (
    DocumentMetadata,
    ExcelLocation,
    NormalizedDocument,
    SourceReference,
    TableElement,
)


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def _non_blank_rows(worksheet: Worksheet) -> list[tuple[int, tuple[Any, ...]]]:
    return [
        (row_number, values)
        for row_number, values in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        )
        if not _is_blank_row(values)
    ]


class ExcelLoader(DocumentLoader):
    """Loads an XLSX workbook into a NormalizedDocument.

    Each non-empty worksheet becomes exactly one TableElement, in workbook
    (worksheet) order. The first non-blank row of a worksheet is its
    header; remaining non-blank rows are its data rows.

    Row numbering convention: ExcelLocation.row_start/row_end are the
    physical (1-indexed) worksheet row numbers of the header row and of
    the final included non-blank row, respectively. Blank rows between
    data rows are omitted from the table body but are NOT used to
    renumber subsequent rows, so row_end reflects the true source
    location even if blank rows were skipped along the way.

    Cell values are preserved as the native Python types openpyxl
    returns (int, float, bool, str, date/datetime, or None) — no
    stringification or custom type inference is performed.
    """

    def load(
        self,
        file_path: str | Path,
        uploaded_by: str | None = None,
    ) -> NormalizedDocument:
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")
        if not path.is_file():
            raise ValueError(f"Excel path is not a file: {path}")

        file_bytes = path.read_bytes()
        file_size = len(file_bytes)
        content_hash = hashlib.sha256(file_bytes).hexdigest()
        document_id = str(uuid.uuid4())

        try:
            workbook = load_workbook(path, data_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to open Excel file {path}: {exc}") from exc

        metadata = DocumentMetadata(
            filename=path.name,
            file_type="excel",
            file_size=file_size,
            content_hash=content_hash,
            uploaded_by=uploaded_by,
            extra={"sheet_names": list(workbook.sheetnames)},
        )

        elements: list[TableElement] = []

        for worksheet in workbook.worksheets:
            non_blank = _non_blank_rows(worksheet)
            if not non_blank:
                continue

            header_row_number, header_values = non_blank[0]
            data_rows = non_blank[1:]
            if not data_rows:
                continue

            columns = ["" if v is None else str(v) for v in header_values]
            rows: list[list[Any]] = [list(values) for _, values in data_rows]
            row_start = header_row_number
            row_end = data_rows[-1][0]

            source = SourceReference(
                document_id=document_id,
                source_type="excel",
                location=ExcelLocation(
                    sheet_name=worksheet.title,
                    row_start=row_start,
                    row_end=row_end,
                ),
            )

            elements.append(
                TableElement(
                    element_id=str(uuid.uuid4()),
                    element_index=len(elements),
                    source=source,
                    section_path=[worksheet.title],
                    columns=columns,
                    rows=rows,
                )
            )

        return NormalizedDocument(
            document_id=document_id,
            metadata=metadata,
            elements=elements,
        )
